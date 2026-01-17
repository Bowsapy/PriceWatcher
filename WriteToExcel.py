import os
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import date

def ExportToExcel():
    conn = sqlite3.connect("prices.db")
    cursor = conn.cursor()

    today_date = date.today().strftime("%d.%m.%Y")


    wb = Workbook()
    cursor.execute("SELECT * from urls")
    products = cursor.fetchall()
    if products:
        wb.remove(wb.active)  # smažeme defaultní sheet

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)

    headers = [
        "Datum",
        "Cena",
    ]

    # 1️⃣ Načteme všechny produkty
    cursor.execute("""
        SELECT id, produkt, heureka_url, moje_cena
        FROM urls
    """)

    products = cursor.fetchall()
    for id, produkt, heureka_url, moje_cena in products:
        sheet_name = f"{id} - {produkt[:25] if produkt else 'Produkt'}"
        ws = wb.create_sheet(title=sheet_name)
        print(id)
        from openpyxl.chart import LineChart, Reference

        # ws je worksheet s daty



        chart = LineChart()
        chart.title = "Historie cen"
        chart.style = 13
        chart.y_axis.title = "Cena"
        chart.x_axis.title = "Datum"
        cursor.execute("""
             SELECT count(id)
             FROM history
         """)
        count = cursor.fetchone()[0] + 4

        # data pro osu Y (sloupec B od řádku 4)
        data = Reference(ws, min_col=2, min_row=4, max_row=count, max_col=2)

        # kategorie pro osu X (sloupec A od řádku 4)
        cats = Reference(ws, min_col=1, min_row=4, max_row=count)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        # přidáme graf do sheetu (např. od sloupce D5)
        ws.add_chart(chart, "D5")

        ws["A1"] = "Produkt:"
        ws["B1"] = produkt

        ws["A2"] = "Heureka URL:"
        ws["B2"] = heureka_url

        ws["A1"].font = Font(bold=True)
        ws["A2"].font = Font(bold=True)
        # hlavičky
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        cursor.execute("""SELECT min_price,max_price,avg_price,act_price from STATISTICS where product_id = ?""",(id,))

        stats = cursor.fetchone()

        if stats:
            min_price, max_price, avg_price, act_price = stats
        else:
            min_price = max_price = avg_price = act_price = None
        ws["D1"] = "Statistiky"
        ws["D2"] = "Min cena"
        ws["D3"] = "Max cena"
        ws["D4"] = "Průměrná cena"
        ws["D5"] = "Aktuální cena"

        ws["E2"] = min_price
        ws["E3"] = max_price
        ws["E4"] = avg_price
        ws["E5"] = act_price


        # 2️⃣ historie pro konkrétní produkt
        cursor.execute("""
            SELECT date, price
            FROM history
            WHERE product_id = ?
            ORDER BY date
        """, (id,))

        history_rows = cursor.fetchall()
        ws.cell(row=1, column=3).value = "Moje cena"
        ws.cell(row=2, column=3).value = moje_cena
        # data
        for row_idx, (hist_date, price) in enumerate(history_rows, start=9):
            ws.cell(row=row_idx, column=1, value=hist_date)
            ws.cell(row=row_idx, column=2, value=price)

        # automatická šířka sloupců
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2


    try:
        wb.save("produkty.xlsx")
        conn.close()
        return True
    except PermissionError:
        return False
ExportToExcel()